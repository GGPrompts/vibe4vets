#!/usr/bin/env python3
"""Export all resources to Excel for local review/editing.

Usage:
    cd backend
    uv run python scripts/export_resources_excel.py

Output: resources_export_YYYYMMDD_HHMMSS.xlsx
"""

import os
import sys
from datetime import datetime

# Add backend to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload
from sqlmodel import Session, select

from app.database import engine
from app.models.resource import Resource


def format_array(arr: list | None) -> str:
    """Format array as comma-separated string for Excel."""
    if not arr:
        return ""
    return ", ".join(str(x) for x in arr)


def export_resources():
    """Export all resources to Excel."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"resources_export_{timestamp}.xlsx"

    print("Fetching resources from database...")

    with Session(engine) as session:
        # Fetch all resources with related data
        stmt = (
            select(Resource)
            .options(
                joinedload(Resource.organization),
                joinedload(Resource.location),
                joinedload(Resource.source),
            )
            .order_by(Resource.created_at.desc())
        )
        resources = session.exec(stmt).unique().all()

        print(f"Found {len(resources)} resources")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resources"

        # Define columns - grouped logically
        columns = [
            # Identity
            ("id", "ID"),
            ("title", "Title"),
            ("status", "Status"),
            # Classification
            ("categories", "Categories"),
            ("subcategories", "Subcategories"),
            ("tags", "Tags"),
            ("scope", "Scope"),
            ("states", "States"),
            # Content
            ("description", "Description"),
            ("summary", "Summary"),
            ("eligibility", "Eligibility"),
            ("how_to_apply", "How to Apply"),
            # Contact
            ("website", "Website"),
            ("phone", "Phone"),
            ("hours", "Hours"),
            ("languages", "Languages"),
            ("cost", "Cost"),
            # Organization
            ("org_name", "Organization"),
            ("org_website", "Org Website"),
            ("org_phones", "Org Phones"),
            # Location
            ("loc_address", "Address"),
            ("loc_city", "City"),
            ("loc_state", "State"),
            ("loc_zip", "ZIP"),
            ("loc_lat", "Latitude"),
            ("loc_lon", "Longitude"),
            # Trust signals
            ("source_name", "Source"),
            ("source_url", "Source URL"),
            ("reliability_score", "Reliability"),
            ("freshness_score", "Freshness"),
            ("last_verified", "Last Verified"),
            ("last_scraped", "Last Scraped"),
            # Link health
            ("link_http_status", "Link HTTP Status"),
            ("link_health_score", "Link Health"),
            ("link_flagged_reason", "Link Flag Reason"),
            ("link_checked_at", "Link Checked"),
            # Timestamps
            ("created_at", "Created"),
            ("updated_at", "Updated"),
            # IDs for reference
            ("organization_id", "Org ID"),
            ("location_id", "Location ID"),
            ("source_id", "Source ID"),
        ]

        # Write header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, (_, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, resource in enumerate(resources, 2):
            org = resource.organization
            loc = resource.location
            src = resource.source

            row_data = {
                # Identity
                "id": str(resource.id),
                "title": resource.title,
                "status": resource.status.value if resource.status else "",
                # Classification
                "categories": format_array(resource.categories),
                "subcategories": format_array(resource.subcategories),
                "tags": format_array(resource.tags),
                "scope": resource.scope.value if resource.scope else "",
                "states": format_array(resource.states),
                # Content
                "description": resource.description,
                "summary": resource.summary,
                "eligibility": resource.eligibility,
                "how_to_apply": resource.how_to_apply,
                # Contact
                "website": resource.website,
                "phone": resource.phone,
                "hours": resource.hours,
                "languages": format_array(resource.languages),
                "cost": resource.cost,
                # Organization
                "org_name": org.name if org else "",
                "org_website": org.website if org else "",
                "org_phones": format_array(org.phones) if org else "",
                # Location
                "loc_address": loc.address if loc else "",
                "loc_city": loc.city if loc else "",
                "loc_state": loc.state if loc else "",
                "loc_zip": loc.zip_code if loc else "",
                "loc_lat": loc.latitude if loc else None,
                "loc_lon": loc.longitude if loc else None,
                # Trust signals
                "source_name": src.name if src else "",
                "source_url": resource.source_url,
                "reliability_score": resource.reliability_score,
                "freshness_score": resource.freshness_score,
                "last_verified": resource.last_verified.isoformat() if resource.last_verified else "",
                "last_scraped": resource.last_scraped.isoformat() if resource.last_scraped else "",
                # Link health
                "link_http_status": resource.link_http_status,
                "link_health_score": resource.link_health_score,
                "link_flagged_reason": resource.link_flagged_reason,
                "link_checked_at": resource.link_checked_at.isoformat() if resource.link_checked_at else "",
                # Timestamps
                "created_at": resource.created_at.isoformat() if resource.created_at else "",
                "updated_at": resource.updated_at.isoformat() if resource.updated_at else "",
                # IDs
                "organization_id": str(resource.organization_id) if resource.organization_id else "",
                "location_id": str(resource.location_id) if resource.location_id else "",
                "source_id": str(resource.source_id) if resource.source_id else "",
            }

            for col_idx, (key, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

            # Progress indicator
            if row_idx % 5000 == 0:
                print(f"  Processed {row_idx - 1} resources...")

        # Auto-size columns (with max width limit)
        for col_idx, (key, _) in enumerate(columns, 1):
            column_letter = get_column_letter(col_idx)
            # Set reasonable widths based on content type
            if key in ("description", "eligibility", "how_to_apply", "summary"):
                ws.column_dimensions[column_letter].width = 50
            elif key in ("title", "org_name", "loc_address"):
                ws.column_dimensions[column_letter].width = 40
            elif key in ("id", "organization_id", "location_id", "source_id"):
                ws.column_dimensions[column_letter].width = 36
            elif key in ("website", "org_website", "source_url"):
                ws.column_dimensions[column_letter].width = 50
            elif key in ("categories", "subcategories", "tags"):
                ws.column_dimensions[column_letter].width = 30
            else:
                ws.column_dimensions[column_letter].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Save
        wb.save(output_file)
        print(f"\nExported to: {output_file}")
        print(f"Total rows: {len(resources)}")
        print(f"Total columns: {len(columns)}")


if __name__ == "__main__":
    export_resources()
